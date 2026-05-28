# caption_app.py
import os
import logging
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
from transformers import AutoProcessor, BlipForConditionalGeneration

# === CONFIGURATION ===
excel_path = 'flickr8k/cleaned_file.xlsx'         # Excel file containing image filenames
sheet_name = 'Sheet1'              # Sheet name
filename_column = 'Filename'      # Column with image filenames
image_folder = 'flickr8k/images/'  # Folder containing images
caption_column_index = 2           # Column index for captions (e.g., column B)
output_file = 'captioned_images_new.xlsx'

# === SETUP LOGGING ===
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%H:%M:%S'
)

# === LOAD MODEL ===
reload_path = os.path.abspath("blip-flickr8k")
logging.info("Loading BLIP model from local path...")
processor = AutoProcessor.from_pretrained(reload_path)
model = BlipForConditionalGeneration.from_pretrained(reload_path)
logging.info("Model loaded successfully.")

# === FUNCTION TO GENERATE CAPTION ===
def caption_image_from_path(image_path: str) -> str:
    image = Image.open(image_path).convert('RGB')
    inputs = processor(images=image, text="a picture of", return_tensors="pt")
    outputs = model.generate(**inputs, max_length=50)
    caption = processor.decode(outputs[0], skip_special_tokens=True)
    return caption

# === LOAD EXCEL ===
logging.info(f"Reading Excel file: {excel_path}")
wb = load_workbook(excel_path)
ws = wb[sheet_name]
df = pd.read_excel(excel_path, sheet_name=sheet_name)

# === CONSTRUCT FULL IMAGE PATHS ===
df[filename_column] = df[filename_column].apply(lambda x: f"{image_folder}{x}")

# === PROCESS EACH IMAGE ===
total = len(df)
logging.info(f"Processing {total} images...")

for idx, row in df.iterrows():
    image_path = row[filename_column]
    log_prefix = f"[{idx + 1}/{total}] {image_path}"

    try:
        caption = caption_image_from_path(image_path)
        logging.info(f"{log_prefix} → Caption generated.")
    except Exception as e:
        caption = f"Error: {str(e)}"
        logging.error(f"{log_prefix} → Failed: {caption}")

    # Write caption to adjacent cell (row offset +2 for header, column B)
    ws.cell(row=idx + 2, column=caption_column_index).value = caption

# === SAVE UPDATED EXCEL ===
wb.save(output_file)
logging.info(f"Captions saved to: {output_file}")