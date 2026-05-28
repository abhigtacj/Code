import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import random

# Load the Excel file
excel_path = 'flickr8k/image_captions.xlsx'
df = pd.read_excel(excel_path)

# Load workbook and select active sheet
wb = load_workbook(excel_path)
ws = wb.active

# Generate a color map for each unique filename
unique_filenames = df['Filename'].unique()
color_map = {}

def random_color():
    return ''.join(random.choices('0123456789ABCDEF', k=6))

for filename in unique_filenames:
    color_map[filename] = PatternFill(start_color=random_color(), end_color=random_color(), fill_type="solid")

# Apply color to rows based on filename
for row_idx, filename in enumerate(df['Filename'], start=2):  # Assuming header is in row 1
    fill = color_map[filename]
    for col_idx in range(1, len(df.columns) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

# Save the modified workbook
wb.save('colored_image_captions.xlsx')